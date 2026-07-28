from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MONTHS = {
    "januari": 1,
    "februari": 2,
    "maret": 3,
    "april": 4,
    "mei": 5,
    "juni": 6,
    "juli": 7,
    "agustus": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "desember": 12,
}

ALIASES = {
    "production": {
        "tahun": ["tahun"],
        "bulan": ["bulan"],
        "commodity": ["nama komoditas", "nama", "jenis tanaman", "komoditas"],
        "luasTanamAkhir": ["luas tan akhir", "luas tanam akhir"],
        "luasPanenHabis": ["luas panen habis"],
        "luasPanenBelumHabis": ["luas panen belum habis"],
        "luasRusak": ["luas rusak"],
        "luasTambahTanam": ["luas tambah tanam"],
        "produksiHabis": ["produksi habis kw", "produksi habis"],
        "produksiBelumHabis": ["produksi belum habis kw", "produksi belum habis"],
        "hargaJual": ["harga jual petani rp kg", "harga jual petani", "harga jual"],
    },
    "weather": {
        "bulan": ["bulan"],
        "suhuMax": ["suhu maks c", "suhu maks"],
        "suhuMin": ["suhu min c", "suhu min"],
        "suhuAvg": ["suhu rata rata c", "suhu rata rata", "suhu rerata"],
        "kecepatanAngin": ["kec angin maks km h", "kec angin maks", "angin maks", "kecepatan angin"],
        "curahHujan": ["curah hujan mm", "curah hujan"],
    },
    "fertilizer": {
        "commodity": ["jenis tanaman", "nama komoditas", "komoditas"],
        "pupuk": ["total pupuk kg", "pupuk kg", "total pupuk"],
        "category": ["kategori"],
    },
}

REQUIRED = {
    "production": ["tahun", "bulan", "commodity", "luasPanenHabis", "luasPanenBelumHabis", "produksiHabis", "produksiBelumHabis"],
    "weather": ["bulan", "suhuMax", "suhuMin", "suhuAvg", "kecepatanAngin", "curahHujan"],
    "fertilizer": ["commodity", "pupuk"],
}


def normalize(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").replace("_", " ")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def month_number(value: Any) -> int:
    if isinstance(value, (int, float)):
        return int(value)
    return MONTHS.get(normalize(value), 0)


def read_rows(path: Path, dataset_type: str | None = None) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))

    wb = load_workbook(path, read_only=True, data_only=True)
    worksheets = wb.worksheets
    if dataset_type == "fertilizer" and worksheets:
        target = next(
            (
                ws
                for ws in worksheets
                if "breakdown pupuk" in normalize(ws.title) or "pupuk" in normalize(ws.title)
            ),
            worksheets[0],
        )
        worksheets = [target]

    candidates: list[dict[str, Any]] = []
    for ws in worksheets:
        raw_rows = list(ws.iter_rows(values_only=True))
        for header_idx, row in enumerate(raw_rows[:8]):
            headers = ["" if cell is None else str(cell) for cell in row]
            if sum(1 for header in headers if header.strip()) < 2:
                continue
            for values in raw_rows[header_idx + 1:]:
                item = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
                item["__sheet"] = ws.title
                candidates.append(item)
            break
    return candidates


def score_dataset(rows: list[dict[str, Any]], dataset_type: str) -> tuple[int, dict[str, str], list[dict[str, Any]], list[str]]:
    mapped, usable_rows, missing = best_rows(rows, dataset_type)
    if dataset_type == "fertilizer":
        usable_rows = [row for row in usable_rows if is_valid_fertilizer_row(row, mapped)]
    score = len(mapped) * 100 + len(usable_rows) - len(missing) * 25
    return score, mapped, usable_rows, missing


def build_mapping(headers: list[str], dataset_type: str) -> dict[str, str]:
    normalized = {normalize(header): header for header in headers}
    mapping: dict[str, str] = {}
    for target, aliases in ALIASES[dataset_type].items():
        for alias in aliases:
            if alias in normalized:
                mapping[target] = normalized[alias]
                break
    return mapping


def detect_year(row: dict[str, Any], mapped: dict[str, str]) -> int:
    year_col = mapped.get("tahun")
    if year_col and to_number(row.get(year_col)):
        return int(to_number(row.get(year_col)) or 0)
    sheet = str(row.get("__sheet", ""))
    match = re.search(r"20\d{2}", sheet)
    return int(match.group(0)) if match else 0


def best_rows(rows: list[dict[str, Any]], dataset_type: str) -> tuple[dict[str, str], list[dict[str, Any]], list[str]]:
    best_mapping: dict[str, str] = {}
    best_rows_list: list[dict[str, Any]] = []
    headers_seen: set[str] = set()

    for row in rows:
        headers = [key for key in row.keys() if key != "__sheet"]
        mapping = build_mapping(headers, dataset_type)
        score = sum(1 for key in REQUIRED[dataset_type] if key in mapping)
        if score > sum(1 for key in REQUIRED[dataset_type] if key in best_mapping):
            best_mapping = mapping
            best_rows_list = []
            headers_seen = set(headers)
        if mapping == best_mapping:
            best_rows_list.append(row)
            headers_seen.update(headers)

    missing = [key for key in REQUIRED[dataset_type] if key not in best_mapping]
    return best_mapping, best_rows_list, missing


def row_commodity(row: dict[str, Any], mapped: dict[str, str]) -> str:
    col = mapped.get("commodity")
    return normalize(row.get(col, "")) if col else ""


def is_valid_fertilizer_row(row: dict[str, Any], mapped: dict[str, str]) -> bool:
    commodity = row_commodity(row, mapped)
    if not commodity or commodity == "jenis tanaman":
        return False
    col = mapped.get("pupuk")
    value = to_number(row.get(col)) if col else None
    return value is not None and value >= 0


def pick_latest(rows: list[dict[str, Any]], mapped: dict[str, str], active_commodity: str, dataset_type: str) -> dict[str, Any] | None:
    active = normalize(active_commodity)
    usable = []
    for row in rows:
        if dataset_type == "fertilizer" and not is_valid_fertilizer_row(row, mapped):
            continue
        if dataset_type in {"production", "fertilizer"} and active and row_commodity(row, mapped) != active:
            continue
        month = month_number(row.get(mapped.get("bulan", "")))
        year = detect_year(row, mapped)
        usable.append((year, month, row))

    if not usable and dataset_type == "fertilizer":
        usable = [(0, 0, row) for row in rows if is_valid_fertilizer_row(row, mapped)]
    if not usable:
        return None
    usable.sort(key=lambda item: (item[0], item[1]))
    return usable[-1][2]


def preview(rows: list[dict[str, Any]], mapped: dict[str, str], limit: int = 5) -> list[dict[str, Any]]:
    output = []
    ordered_keys = [
        "tahun",
        "bulan",
        "commodity",
        "luasTanamAkhir",
        "luasPanenHabis",
        "luasPanenBelumHabis",
        "luasRusak",
        "luasTambahTanam",
        "produksiHabis",
        "produksiBelumHabis",
        "hargaJual",
        "suhuMax",
        "suhuMin",
        "suhuAvg",
        "curahHujan",
        "kecepatanAngin",
        "pupuk",
        "category",
    ]
    mapped_cols = [mapped[key] for key in ordered_keys if key in mapped]
    for row in rows[:limit]:
        item = {}
        for key in mapped_cols:
            if key in row and row[key] is not None:
                item[key] = row[key]
        if item:
            output.append(item)
    return output


def form_patch(row: dict[str, Any] | None, mapped: dict[str, str], dataset_type: str) -> dict[str, float]:
    if not row:
        return {}
    patch: dict[str, float] = {}
    if dataset_type == "production":
        for key in ["luasTanamAkhir", "luasPanenHabis", "luasPanenBelumHabis", "luasRusak", "luasTambahTanam", "hargaJual"]:
            col = mapped.get(key)
            value = to_number(row.get(col)) if col else None
            if value is not None:
                patch[key] = value
        for key in ["produksiHabis", "produksiBelumHabis"]:
            col = mapped.get(key)
            value = to_number(row.get(col)) if col else None
            if value is not None:
                patch[key] = value * 100
    elif dataset_type == "weather":
        for key in ["suhuMax", "suhuMin", "suhuAvg", "curahHujan"]:
            col = mapped.get(key)
            value = to_number(row.get(col)) if col else None
            if value is not None:
                patch[key] = value
        col = mapped.get("kecepatanAngin")
        value = to_number(row.get(col)) if col else None
        if value is not None:
            patch["kecepatanAngin"] = value / 3.6 if value > 15 else value
    elif dataset_type == "fertilizer":
        col = mapped.get("pupuk")
        value = to_number(row.get(col)) if col else None
        if value is not None:
            patch["pupuk"] = value
    return patch


def commodity_options(rows: list[dict[str, Any]], mapped: dict[str, str], dataset_type: str) -> list[dict[str, Any]]:
    if dataset_type != "production" or "commodity" not in mapped:
        return []

    commodity_col = mapped["commodity"]
    latest_by_name: dict[str, tuple[int, int, dict[str, Any]]] = {}
    history_by_name: dict[str, list[tuple[int, int, dict[str, Any]]]] = {}
    for row in rows:
        name = str(row.get(commodity_col) or "").strip()
        if not name:
            continue
        key = normalize(name)
        year = detect_year(row, mapped)
        month = month_number(row.get(mapped.get("bulan", "")))
        history_by_name.setdefault(key, []).append((year, month, row))
        current = latest_by_name.get(key)
        if current is None or (year, month) >= (current[0], current[1]):
            latest_by_name[key] = (year, month, row)

    output = []
    for _, _, row in latest_by_name.values():
        name = str(row.get(commodity_col) or "").strip()
        produksihabis = to_number(row.get(mapped.get("produksiHabis", ""))) or 0
        produksi_belum = to_number(row.get(mapped.get("produksiBelumHabis", ""))) or 0
        harga = to_number(row.get(mapped.get("hargaJual", ""))) or 0
        history = []
        previous_actual = 0.0
        for year, month, history_row in sorted(history_by_name.get(normalize(name), []), key=lambda item: (item[0], item[1])):
            actual = (
                (to_number(history_row.get(mapped.get("produksiHabis", ""))) or 0)
                + (to_number(history_row.get(mapped.get("produksiBelumHabis", ""))) or 0)
            ) * 100
            if actual <= 0 and previous_actual <= 0:
                predicted = 0
            elif previous_actual > 0:
                predicted = round((actual * 0.7) + (previous_actual * 0.3))
            else:
                predicted = round(actual * 0.96)
            previous_actual = actual
            label = f"{str(history_row.get(mapped.get('bulan', ''), '')).strip()[:3]} {year}" if year else str(history_row.get(mapped.get("bulan", ""), "")).strip()
            history.append({
                "month": label,
                "actual": round(actual),
                "predicted": round(predicted),
            })
        output.append({
            "name": name,
            "category": "Buah-buahan" if normalize(name) in {"stroberi", "melon", "semangka"} else "Sayuran",
            "price": harga,
            "productionKg": (produksihabis + produksi_belum) * 100,
            "formPatch": form_patch(row, mapped, "production"),
            "history": history[-12:],
        })
    return sorted(output, key=lambda item: item["name"])


def main() -> None:
    path = Path(sys.argv[1])
    dataset_type = sys.argv[2]
    active_commodity = sys.argv[3] if len(sys.argv) > 3 else ""

    raw_rows = read_rows(path)
    kandidat = [dataset_type] + [t for t in REQUIRED.keys() if t != dataset_type]
    hasil = [score_dataset(raw_rows, t) + (t,) for t in kandidat]
    hasil.sort(key=lambda item: item[0], reverse=True)

    _, mapped, usable_rows, missing, detected_type = hasil[0]
    rows = read_rows(path, detected_type)
    if rows and rows is not raw_rows:
        _, mapped, usable_rows, missing = score_dataset(rows, detected_type)

    latest = pick_latest(usable_rows, mapped, active_commodity, detected_type)
    patch = form_patch(latest, mapped, detected_type)

    result = {
        "valid": not missing and bool(usable_rows),
        "datasetType": detected_type,
        "rowsCount": len(usable_rows),
        "missingColumns": missing,
        "mappedColumns": mapped,
        "previewRows": preview(usable_rows, mapped, limit=25),
        "commodityOptions": commodity_options(usable_rows, mapped, detected_type),
        "appliedPatch": patch,
        "appliedFrom": latest,
        "message": "Dataset valid dan nilai terbaru sudah siap dipakai." if not missing else "Kolom wajib belum lengkap.",
    }
    print(json.dumps(result, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
